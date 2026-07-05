"""Фикс-спецификации бланка + «достоверная» отрисовка листа Excel.

КОЛОНКИ таблиц НЕ детектируются, а заданы здесь (на этих сканах вертикальные
рули бледные — геометрия колонок ненадёжна). write_spec рисует двухуровневую
шапку с merge_cells, ряд номеров граф (где он есть), рамки, ширины и поворот
вертикальных заголовков — структура верна by construction.

Метки строк у podschet/pokazateli/kontrol ИЗВЕСТНЫ (row_labels) — фиксируем их
из спеки, от движка берём только значения (устойчивее на бледных мини-таблицах).

Подписи печатной шапки местами бледные — часть меток предварительна; уточнить
при разметке эталона (структура/число граф достоверны).
"""
from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_THIN = Side(style="thin", color="808080")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
HEAD_FILL = PatternFill("solid", fgColor="D9E1F2")
DITTO_FILL = PatternFill("solid", fgColor="FFF2CC")    # жёлтый — протянутый «-ll-»
LOWCONF_FILL = PatternFill("solid", fgColor="FCE4D6")  # оранжевый — needs_review
LABEL_FILL = PatternFill("solid", fgColor="EDEDED")    # серый — фикс-метки строк
BOLD = Font(bold=True, size=9)
TITLE_FONT = Font(bold=True, size=12)
VERT = Alignment(textRotation=90, horizontal="center", vertical="center", wrap_text=True)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LOWCONF = 0.55


@dataclass
class Spec:
    kind: str
    title: str
    columns: list[tuple]                 # ("single",label,w,rot) | ("group",label,[(sub,w,rot)...])
    graph_nums: list[int] | None = None  # ряд печатных номеров граф (если есть)
    row_labels: list[str] | None = None  # фикс-метки строк в колонке 1
    detect: dict = field(default_factory=dict)
    # Пиксельные доли ширины колонок (границы X-полос для ГЕОМЕТРИЧЕСКОЙ раскладки),
    # относительно ширины региона объекта. Если None — берутся из widths.
    # Калибруются по эталонному бланку (детекция вертикалей / замер).
    col_fracs: list[float] | None = None
    header_rows: int = 2                  # 1 (одноуровневая) | 2 (двухуровневая шапка)
    units: list[str] | None = None        # фикс «единица измерения» по строкам (подсчёт)
    # индексы колонок, которые движок ЗАПОЛНЯЕТ (для labelled-таблиц); прочие — из спеки
    fill_cols: list[int] | None = None
    # колонка-подпись, сливаемая на всю высоту (E1:E7 у подсчёта): индекс или None
    signature_col: int | None = None
    # раскладка значения по X-ЦЕНТРУ слова Yandex (без покадрового кропа) —
    # для таблиц, где значение = отдельное число/код (не режем на подстроки)
    word_placement: bool = False

    @property
    def ncols(self) -> int:
        n = 0
        for seg in self.columns:
            n += 1 if seg[0] == "single" else len(seg[2])
        return n

    @property
    def widths(self) -> list[int]:
        w: list[int] = []
        for seg in self.columns:
            if seg[0] == "single":
                w.append(seg[2])
            else:
                w.extend(ww for (_, ww, _) in seg[2])
        return w


# --------------------------- СПЕЦИФИКАЦИИ ---------------------------------
JOURNAL_LEFT = Spec(
    kind="journal_left",
    title="Буровой журнал — графы 1–8 (описание разреза)",
    columns=[
        ("single", "№ проходок (проб)", 5, True),
        ("group", "Глубина, м", [("скважины", 5, True), ("обсадки трубами", 5, True)]),
        ("single", "Литологический разрез", 5, True),
        ("single", "ОПИСАНИЕ РАЗРЕЗА", 36, False),
        ("single", "Отметка о таликах, мерзлоте и водоносности", 6, True),
        ("single", "Категория породы", 6, True),
        ("single", "Объём извлечённой породы, см³", 8, True),
    ],
    graph_nums=list(range(1, 9)),
    detect={"cols": 8, "keywords": ["описание", "проходок", "глубина", "разрез"]},
    # замерено по эталону (вертикальные рули): № узкая, ОПИСАНИЕ ~0.40 в центре
    col_fracs=[0.049, 0.076, 0.082, 0.081, 0.401, 0.107, 0.089, 0.115],
)

SAMPLING_RIGHT = Spec(
    kind="sampling_right",
    title="Опробование — графы 9–18",
    columns=[
        ("single", "Объём породы теоретический, см³", 8, True),
        ("single", "Диаметр скважины по манометру", 6, True),
        ("group", "Результат опробования, мг",
         [("визуально", 5, True), ("лаборатория, вес", 5, True)]),
        ("single", "Вес металла, получен. от дополнит. промывки", 6, True),
        ("single", "Вес металла, привяз. к подсчёту", 6, True),
        ("group", "Процент", [("каменистости", 5, True), ("ледянистости", 5, True)]),
        ("single", "Среднее содержание по промывке", 6, True),
        ("single", "Содержание металла на куб. м с учётом приборности", 8, True),
    ],
    graph_nums=list(range(9, 19)),
    detect={"cols": 10, "keywords": ["опроб", "теоретич", "процент", "содержание"]},
    # замерено по словам (xc): коды г11~0.26, значения г12~0.38, числа
    # г16~0.767, г17~0.90; границы графов 16/17/18 = 0.70/0.82/0.93
    col_fracs=[0.10, 0.10, 0.10, 0.10, 0.10, 0.10, 0.10, 0.12, 0.11, 0.07],
    # покадровый апскейл-OCR по калиброванным X-полосам: одна графа = одно число
    # (читает бледные 190/48, разделяет слипшееся «19 54»)
)

SAMORODKI = Spec(
    kind="samorodki",
    title="Характеристика самородков",
    columns=[
        ("single", "№ проходок (проб)", 6, True),
        ("single", "Количество и вес самородков", 16, False),
        ("single", "Размер в трёх измерениях", 12, False),
        ("single", "Окатанность, включая налёты, схематическая зарисовка и др.", 24, False),
    ],
    graph_nums=None, header_rows=1,
    detect={"cols": 4, "keywords": ["самородк", "окатан", "размер"]},
)

PODSCHET = Spec(
    kind="podschet",
    title="Пробность золота: содержание олова проц.",
    columns=[
        ("single", "Наименование показателей", 28, False),
        ("group", "Лимитность",
         [("единица измерения", 10, False), ("на пласт.", 9, False), ("на массу", 9, False)]),
        ("single", "Подсчёт произвёл", 18, False),
    ],
    graph_nums=None, header_rows=2,
    row_labels=["Глубина выемки", "Мощность торфов", "Мощность песков",
                "Среднее содержание", "Вертикальный запас"],
    units=["м", "-ll-", "-ll-", "мг-м куб.", "мг-м куб."],  # фикс графа «единица изм.» (повтор = ditto)
    fill_cols=[2, 3],          # движок заполняет ТОЛЬКО «на пласт» и «на массу» (0-based)
    signature_col=5,           # «Подсчёт произвёл» = колонка E (1-based) -> merge E1:E7
    # замерено по словам: пласт xc~0.49, масса xc~0.55, подпись xc>0.60
    col_fracs=[0.37, 0.06, 0.095, 0.075, 0.40],
    word_placement=True,       # значения = отдельные числа -> по X-центру, подпись отсекается
    detect={"cols": 5, "keywords": ["подсч", "наименование показател", "пласт"]},
)

POKAZATELI = Spec(
    kind="pokazateli",
    title="ПОКАЗАТЕЛИ",
    columns=[
        ("single", "ПОКАЗАТЕЛИ", 34, False),
        ("single", "На пласт.", 10, False),
        ("single", "На выем. мощ", 10, False),
        ("single", "На массу", 10, False),
    ],
    graph_nums=None, header_rows=1, fill_cols=[1, 2, 3],
    row_labels=["Общий вес металла по скважине, мг", "Вес металла введенного в подсчёт, мг",
                "Площадь скважин для теорет. объема, кв. см",
                "Средний коэффициент по объему"],
    detect={"cols": 4, "keywords": ["показател", "на пласт", "на массу"]},
)

KONTROL = Spec(
    kind="kontrol",
    title="Контроль",
    columns=[
        ("single", "Дата контроля", 14, False),
        ("single", "Что контролируется", 28, False),
        ("single", "Вес металла", 12, False),
    ],
    graph_nums=None, header_rows=1,
    detect={"cols": 3, "keywords": ["контрол", "что контролир", "вес металла"]},
)

SPECS: dict[str, Spec] = {
    s.kind: s for s in (JOURNAL_LEFT, SAMPLING_RIGHT, SAMORODKI, PODSCHET, POKAZATELI, KONTROL)
}


# --------------------------- detect_kind ---------------------------------
def _looks_like_header(row: list[str]) -> bool:
    first = (row[0] if row else "").strip()
    return not (first.isdigit() or first in ("", "1"))


def detect_kind(table: dict[str, Any]) -> str | None:
    """Тип таблицы по числу колонок + ключевым словам печатной шапки."""
    cols = table.get("cols") or 0
    grid = table.get("grid") or []
    head = " ".join(" ".join(r) for r in grid[:2]).lower() if grid else ""
    best = None
    for spec in SPECS.values():
        if spec.detect.get("cols") != cols:
            continue
        hits = sum(1 for kw in spec.detect.get("keywords", []) if kw in head)
        if best is None or hits > best[1]:
            best = (spec.kind, hits)
    return best[0] if best else None


# --------------------------- отрисовка -----------------------------------
# Шапка на строках 1..header_rows (как в эталоне заказчика — так совпадают merges).
def _write_header(ws, spec: Spec) -> None:
    for i, w in enumerate(spec.widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    if spec.header_rows == 1:
        c = 1
        for seg in spec.columns:                 # у малых таблиц только single
            cell = ws.cell(1, c, seg[1])
            cell.font = BOLD; cell.fill = HEAD_FILL; cell.border = BORDER
            cell.alignment = VERT if (len(seg) > 3 and seg[3]) else CTR
            c += 1
        ws.row_dimensions[1].height = 46
        return
    # header_rows == 2
    c = 1
    for seg in spec.columns:
        if seg[0] == "single":
            _, label, _, rot = seg
            is_title0 = spec.kind == "podschet" and c == 1        # A1=title, A2=label (без merge)
            is_sig = spec.signature_col is not None and c == spec.signature_col
            if is_title0:
                for rr, txt in ((1, spec.title), (2, label)):
                    cc = ws.cell(rr, c, txt)
                    cc.font = BOLD; cc.fill = HEAD_FILL; cc.border = BORDER; cc.alignment = CTR
            elif is_sig:
                cc = ws.cell(1, c, label)         # merge на всю высоту — позже
                cc.font = BOLD; cc.fill = HEAD_FILL; cc.alignment = CTR
            else:
                cell = ws.cell(1, c, label)
                ws.merge_cells(start_row=1, start_column=c, end_row=2, end_column=c)
                cell.font = BOLD; cell.fill = HEAD_FILL; cell.alignment = VERT if rot else CTR
                for rr in (1, 2):
                    ws.cell(rr, c).border = BORDER; ws.cell(rr, c).fill = HEAD_FILL
            c += 1
        else:
            _, glabel, subs = seg
            n = len(subs)
            g = ws.cell(1, c, glabel)
            ws.merge_cells(start_row=1, start_column=c, end_row=1, end_column=c + n - 1)
            g.font = BOLD; g.fill = HEAD_FILL; g.alignment = CTR
            for j, (sub, _, rot) in enumerate(subs):
                sc = ws.cell(2, c + j, sub)
                sc.font = BOLD; sc.fill = HEAD_FILL; sc.alignment = VERT if rot else CTR
                sc.border = BORDER
            for jj in range(n):
                ws.cell(1, c + jj).border = BORDER
            c += n
    ws.row_dimensions[1].height = 110
    ws.row_dimensions[2].height = 64


def write_spec(ws, spec: Spec, table: dict[str, Any], extra_blank: int = 2) -> None:
    grid = table.get("grid") or []
    conf = table.get("conf") or []
    expanded = table.get("expanded") or []
    ncols = spec.ncols
    _write_header(ws, spec)

    data_top = spec.header_rows + 1
    if spec.graph_nums:
        for i, num in enumerate(spec.graph_nums[:ncols], 1):
            cell = ws.cell(data_top, i, num)
            cell.font = BOLD; cell.alignment = CTR; cell.border = BORDER
        ws.row_dimensions[data_top].height = 14
        data_top += 1

    wrap_cols = {i for i, w in enumerate(spec.widths) if w >= 20}
    if spec.row_labels:
        n_data = _write_labelled(ws, spec, grid, conf, data_top, wrap_cols)
    else:
        start = 1 if (grid and _looks_like_header(grid[0])) else 0
        n_data = _write_freeform(ws, spec, grid[start:], conf, expanded, start,
                                 data_top, wrap_cols, extra_blank)
    # подпись «Подсчёт произвёл» -> слияние на всю высоту (E1:E7)
    if spec.signature_col is not None and n_data:
        last = data_top + n_data - 1
        ws.merge_cells(start_row=1, start_column=spec.signature_col,
                       end_row=last, end_column=spec.signature_col)
    ws.freeze_panes = ws.cell(data_top, 1)


def _write_labelled(ws, spec: Spec, grid, conf, dtop, wrap_cols) -> int:
    """col0 = фикс-метка; «единица измерения» = из спеки; значения — только в
    fill_cols; подпись — пусто. Движок метки/единицы НЕ перезаписывает."""
    ncols = spec.ncols
    fill = set(spec.fill_cols or range(1, ncols))
    for i, label in enumerate(spec.row_labels):
        r = dtop + i
        lc = ws.cell(r, 1, label)
        lc.border = BORDER; lc.alignment = LEFT; lc.fill = LABEL_FILL
        for ci in range(1, ncols):
            if spec.units and ci == 1:
                val = spec.units[i] if i < len(spec.units) else ""
            elif ci in fill:
                val = grid[i][ci] if i < len(grid) and ci < len(grid[i]) else ""
            else:
                val = ""                          # подпись/непубличные графы — пусто
            cell = ws.cell(r, ci + 1, val)
            cell.border = BORDER
            cell.alignment = LEFT if ci in wrap_cols else CTR
            if spec.units and ci == 1:
                cell.fill = LABEL_FILL
            cf = _conf_at(conf, i, ci)
            if ci in fill and cf is not None and float(cf) < LOWCONF:
                cell.fill = LOWCONF_FILL
    return len(spec.row_labels)


def _write_freeform(ws, spec, data, conf, expanded, start, dtop, wrap_cols, extra_blank) -> int:
    ncols = spec.ncols
    for di, row in enumerate(data):
        r = dtop + di
        gi = start + di
        for ci in range(ncols):
            val = row[ci] if ci < len(row) else ""
            cell = ws.cell(r, ci + 1, val)
            cell.border = BORDER
            cell.alignment = LEFT if ci in wrap_cols else CTR
            if gi < len(expanded) and ci < len(expanded[gi]) and expanded[gi][ci]:
                cell.fill = DITTO_FILL
            cf = _conf_at(conf, gi, ci)
            if cf is not None and float(cf) < LOWCONF:
                cell.fill = LOWCONF_FILL
    for k in range(extra_blank):
        r = dtop + len(data) + k
        for ci in range(ncols):
            ws.cell(r, ci + 1).border = BORDER
    return len(data)


def _conf_at(conf, r, c):
    if r < len(conf) and c < len(conf[r]):
        return conf[r][c]
    return None


def _similar(a: str, b: str) -> bool:
    na = "".join(ch for ch in (a or "").lower() if ch.isalnum())
    nb = "".join(ch for ch in (b or "").lower() if ch.isalnum())
    return bool(na) and na[:6] == nb[:6]


def write_table_for_kind(ws, kind: str, table: dict[str, Any]) -> bool:
    spec = SPECS.get(kind)
    if spec is None:
        return False
    write_spec(ws, spec, table)
    return True


def write_table(ws, table: dict[str, Any]) -> bool:
    kind = detect_kind(table)
    return write_table_for_kind(ws, kind, table) if kind else False
