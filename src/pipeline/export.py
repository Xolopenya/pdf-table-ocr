"""Экспорт результатов в .xlsx (+ .json). Один объект = один лист.

Таблицы рисует forms_layout.write_table_for_kind (шапка со слияниями «как в
бланке»). Формы — key-value листами. Служебный лист `_meta`: тип объекта, bbox,
источник текста (qwen/yandex/hybrid), обе версии текста в гибриде, conf,
needs_review. Имена листов ≤31 симв., уникальные.
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from . import forms_layout as fl

_THIN = Side(style="thin", color="999999")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
HEAD_FILL = PatternFill("solid", fgColor="D9E1F2")
REVIEW_FILL = PatternFill("solid", fgColor="FCE4D6")
BOLD = Font(bold=True)
TITLE = Font(bold=True, size=12)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _unique(name: str, used: set[str]) -> str:
    for ch in r'[]:*?/\\':
        name = name.replace(ch, "_")
    name = name[:31]
    if name in used:
        # НИКАКИХ молчаливых «_2»: коллизия имён листа = ошибка идентификации
        raise ValueError(f"Коллизия имени листа '{name}' — вероятно дублируется "
                         f"номер скважины. Проверьте borehole_numbers / позиции.")
    used.add(name)
    return name


def _write_form_sheet(ws, form: dict[str, Any]) -> None:
    ws.cell(1, 1, form["title"]).font = TITLE
    r = 3
    for h, txt in enumerate(("Поле", "Значение", "Источник")):
        c = ws.cell(r, h + 1, txt)
        c.font = BOLD
        c.fill = HEAD_FILL
        c.border = BORDER
    r += 1
    for f in form["fields"]:
        kc = ws.cell(r, 1, f["key"]); kc.alignment = WRAP; kc.border = BORDER
        vc = ws.cell(r, 2, f["value"]); vc.alignment = WRAP; vc.border = BORDER
        sc = ws.cell(r, 3, f.get("source", "")); sc.alignment = CTR; sc.border = BORDER
        if f.get("needs_review"):
            vc.fill = REVIEW_FILL
        r += 1
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 12
    ws.freeze_panes = "A4"


def _write_meta_sheet(ws, meta: list[dict[str, Any]]) -> None:
    cols = ["sheet", "obj_type", "r", "c", "rs", "cs", "bbox",
            "text_source", "qwen_text", "yandex_text", "conf", "needs_review", "text"]
    for i, name in enumerate(cols, 1):
        c = ws.cell(1, i, name); c.font = BOLD; c.fill = HEAD_FILL
    for ri, m in enumerate(meta, 2):
        vals = [m.get("sheet"), m.get("obj_type"), m.get("r"), m.get("c"),
                m.get("rs", 1), m.get("cs", 1),
                json.dumps(m.get("bbox")) if m.get("bbox") else "",
                m.get("text_source"), m.get("qwen_text"), m.get("yandex_text"),
                m.get("conf"), m.get("needs_review"), m.get("text")]
        for ci, v in enumerate(vals, 1):
            ws.cell(ri, ci, v)
    ws.freeze_panes = "A2"


def build_workbook(sheets: list[dict[str, Any]], meta: list[dict[str, Any]],
                   write_meta: bool = True) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    for sh in sheets:
        name = _unique(sh["name"], used)
        ws = wb.create_sheet(name)
        if sh.get("form") is not None:
            _write_form_sheet(ws, sh["form"])
        else:
            table = sh["table"]
            if not fl.write_table_for_kind(ws, sh["kind"], table):
                # запасной вариант — простая сетка
                _write_plain(ws, table)
    if write_meta and meta:
        _write_meta_sheet(wb.create_sheet("_meta"), meta)
    return wb


def _write_plain(ws, table: dict[str, Any]) -> None:
    grid = table.get("grid") or []
    for i, row in enumerate(grid, 1):
        for j, v in enumerate(row, 1):
            cell = ws.cell(i, j, v)
            cell.border = BORDER
            cell.alignment = CTR


def export_document(sheets, meta, out_base: str | Path, *, source: str, engine: str,
                    write_meta: bool = True) -> dict[str, str]:
    out_base = Path(out_base)
    out_base.parent.mkdir(parents=True, exist_ok=True)
    xlsx = out_base.with_suffix(".xlsx")
    js = out_base.with_suffix(".json")
    wb = build_workbook(sheets, meta, write_meta=write_meta)
    try:
        wb.save(xlsx)
    except PermissionError:
        # файл открыт (напр., в Excel) -> сохраняем рядом с суффиксом времени
        import time
        alt = out_base.with_name(out_base.name + f"_{time.strftime('%H%M%S')}").with_suffix(".xlsx")
        wb.save(alt)
        xlsx = alt
    payload = {"source": source, "engine": engine,
               "sheets": [{"name": s["name"], "kind": s["kind"],
                           "table": s.get("table"), "form": s.get("form")} for s in sheets],
               "meta": meta}
    js.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"xlsx": str(xlsx), "json": str(js)}
